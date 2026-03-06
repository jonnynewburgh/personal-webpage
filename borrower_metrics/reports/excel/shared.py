"""
Shared Excel helpers: palette, cell styling, summary sheet builder.
"""
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
NAVY_HEX  = "1B3A6B"
TEAL_HEX  = "2A7F8F"
GOLD_HEX  = "C8952A"
LIGHT_HEX = "E8EDF4"
GRAY_HEX  = "6C757D"
GREEN_HEX = "2E7D4F"
RED_HEX   = "C0392B"
WHITE_HEX = "FFFFFF"

ORG_TYPE_LABEL = {
    "charter_school": "Charter School",
    "fqhc":           "Federally Qualified Health Center",
    "health_center":  "Federally Qualified Health Center",
    "early_care":     "Early Care & Education Program",
    "nonprofit":      "Nonprofit Organization",
}


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(name="Calibri", size=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, cols, fill_hex=NAVY_HEX, font_color=WHITE_HEX, size=10):
    for col, val in enumerate(cols, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(fill_hex)
        c.font      = _font(bold=True, color=font_color, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()


def _data_row(ws, row, vals, alt=False):
    bg = LIGHT_HEX if alt else WHITE_HEX
    for col, val in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.border    = _border("hair")
        c.alignment = Alignment(horizontal="center", vertical="center")


def _pct(val):
    return f"{val:.1f}%" if val is not None else "N/A"


def _section_header(ws, row, label, n_cols=6, fill_hex=TEAL_HEX):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font      = _font(bold=True, color=WHITE_HEX, size=10)
    c.fill      = _fill(fill_hex)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _embed_chart(ws, buf, anchor, width=400, height=240):
    img = XLImage(buf)
    img.width, img.height = width, height
    ws.add_image(img, anchor)


def build_summary_sheet(wb, profile) -> None:
    """
    Build the Summary sheet — always present, content adapts to org type.
    """
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = profile.name
    c.font      = _font(size=16, bold=True, color=NAVY_HEX)
    c.fill      = _fill(LIGHT_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Core metadata
    meta = [
        ("Organization Type", ORG_TYPE_LABEL.get(profile.org_type, profile.org_type)),
        ("Location",          profile.location),
        ("EIN",               profile.ein or "—"),
        ("Year Founded",      profile.year_founded or "—"),
        ("Website",           profile.website or "—"),
        ("Report Date",       profile.report_date or date.today().strftime("%B %d, %Y")),
        ("Prepared By",       profile.prepared_by or "—"),
    ]

    row = 2
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=NAVY_HEX)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Org-type-specific key facts
    row += 1
    _section_header(ws, row, "KEY FACTS", n_cols=6, fill_hex=NAVY_HEX)
    row += 1
    kf = _key_facts(profile)
    for label, val in kf:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=NAVY_HEX)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Loan block
    if profile.loan:
        row += 1
        _section_header(ws, row, "PROPOSED FINANCING", n_cols=6, fill_hex=NAVY_HEX)
        row += 1
        loan_meta = [
            ("Loan Amount",   f"${profile.loan.amount:,.0f}"),
            ("Purpose",       profile.loan.purpose),
            ("Structure",     profile.loan.financing_type),
            ("Term",          f"{profile.loan.proposed_term_years} years"
                              if profile.loan.proposed_term_years else "—"),
            ("Collateral",    profile.loan.collateral or "—"),
        ]
        for label, val in loan_meta:
            ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=NAVY_HEX)
            ws.cell(row=row, column=2, value=val)
            row += 1

    # Analyst notes
    if profile.analyst_notes:
        row += 1
        _section_header(ws, row, "ANALYST NOTES", n_cols=6, fill_hex=TEAL_HEX)
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=profile.analyst_notes)
        c.font      = _font(italic=True)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50

    _set_col_widths(ws, [22, 40, 15, 15, 15, 15])


def _key_facts(profile):
    """Return list of (label, value) tuples specific to the org type."""
    from borrower_metrics.models import (
        CharterSchoolProfile, FQHCProfile, EarlyCareProfile, NonprofitProfile
    )
    facts = []

    if isinstance(profile, CharterSchoolProfile):
        if profile.authorizer:
            facts.append(("Authorizer", profile.authorizer))
        if profile.grade_span:
            facts.append(("Grade Span", profile.grade_span))
        if profile.management_org:
            facts.append((profile.management_org_type or "Mgmt Org", profile.management_org))
        if profile.facility_status:
            facts.append(("Facility Status", profile.facility_status))
        if profile.per_pupil_revenue:
            facts.append(("Per-Pupil Revenue", f"${profile.per_pupil_revenue:,.0f}"
                          + (f" ({profile.per_pupil_revenue_year})"
                             if profile.per_pupil_revenue_year else "")))
        if profile.free_reduced_lunch_pct is not None:
            facts.append(("FRL %", _pct(profile.free_reduced_lunch_pct)))
        if profile.english_learners_pct is not None:
            facts.append(("ELL %", _pct(profile.english_learners_pct)))
        if profile.special_education_pct is not None:
            facts.append(("SPED %", _pct(profile.special_education_pct)))

    elif isinstance(profile, FQHCProfile):
        if profile.number_of_sites:
            facts.append(("Number of Sites", profile.number_of_sites))
        if profile.ftca_deemed is not None:
            facts.append(("FTCA Deemed", "Yes" if profile.ftca_deemed else "No"))
        if profile.hrsa_grant:
            g = profile.hrsa_grant
            if g.grant_amount_annual:
                facts.append(("Annual 330 Grant", f"${g.grant_amount_annual:,.0f}"))
            if g.award_period_start and g.award_period_end:
                facts.append(("Award Period", f"{g.award_period_start}–{g.award_period_end}"))
            if g.award_number:
                facts.append(("Award Number", g.award_number))
            facts.append(("HRSA Type",
                           "Look-Alike" if g.look_alike else "Section 330 Grantee"))
        if profile.last_osa_review_year:
            facts.append(("Last OSV", profile.last_osa_review_year +
                          (f" — {profile.osa_outcome}" if profile.osa_outcome else "")))
        if profile.service_area:
            facts.append(("Service Area", profile.service_area))

    elif isinstance(profile, EarlyCareProfile):
        if profile.qris_rating:
            rating = profile.qris_rating
            if profile.qris_framework:
                rating += f" ({profile.qris_framework})"
            if profile.qris_rating_year:
                rating += f" — {profile.qris_rating_year}"
            facts.append(("QRIS Rating", rating))
        flags = []
        if profile.head_start_grantee:
            flags.append("Head Start")
        if profile.early_head_start_grantee:
            flags.append("Early Head Start")
        if flags:
            facts.append(("Federal Programs", ", ".join(flags)))
        if profile.licensed_capacity_total:
            facts.append(("Licensed Capacity", profile.licensed_capacity_total))
        if profile.number_of_classrooms:
            facts.append(("Classrooms", profile.number_of_classrooms))
        if profile.subsidized_enrollment_pct is not None:
            facts.append(("Subsidized Enrollment", _pct(profile.subsidized_enrollment_pct)))
        if profile.income_eligible_pct is not None:
            facts.append(("Income Eligible", _pct(profile.income_eligible_pct)))
        if profile.last_monitoring_year:
            facts.append(("Last Monitoring",
                           profile.last_monitoring_year +
                           (f" — {profile.monitoring_outcome}"
                            if profile.monitoring_outcome else "")))

    elif isinstance(profile, NonprofitProfile):
        if profile.primary_program_area:
            facts.append(("Program Area", profile.primary_program_area))
        if profile.clients_served_annually:
            facts.append(("Clients Served",
                           f"{profile.clients_served_annually:,}" +
                           (f" ({profile.clients_served_year})"
                            if profile.clients_served_year else "")))
        if profile.operating_reserve_months is not None:
            facts.append(("Operating Reserve",
                           f"{profile.operating_reserve_months:.1f} months" +
                           (f" ({profile.operating_reserve_year})"
                            if profile.operating_reserve_year else "")))
        if profile.last_audit_year:
            facts.append(("Last Audit",
                           profile.last_audit_year +
                           (f" — {profile.audit_outcome}" if profile.audit_outcome else "")))
        if profile.irs_form_990_year:
            facts.append(("Form 990 Year", profile.irs_form_990_year))

    return facts
