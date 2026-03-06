"""
Excel workbook generator using openpyxl.
Sheets:
  1. Summary       – mirrors the one-pager at a glance
  2. Enrollment    – year-by-year table + embedded bar chart
  3. Demographics  – breakdown table + embedded chart
  4. Academics     – proficiency / growth by year + embedded chart
  5. Charter Log   – chronological event log
  6. Accountability– ratings and flags
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .models import BorrowerProfile
from . import charts

# ── Palette ───────────────────────────────────────────────────────────────
NAVY_HEX  = "1B3A6B"
TEAL_HEX  = "2A7F8F"
GOLD_HEX  = "C8952A"
LIGHT_HEX = "E8EDF4"
GRAY_HEX  = "6C757D"
GREEN_HEX = "2E7D4F"
RED_HEX   = "C0392B"
WHITE_HEX = "FFFFFF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(name="Calibri", size=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, cols, fill_hex=NAVY_HEX, font_color=WHITE_HEX, size=10):
    for col, val in enumerate(cols, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill  = _fill(fill_hex)
        c.font  = _font(bold=True, color=font_color, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()

def _data_row(ws, row, vals, alt=False):
    bg = LIGHT_HEX if alt else WHITE_HEX
    for col, val in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = _fill(bg)
        c.border = _border("hair")
        c.alignment = Alignment(horizontal="center", vertical="center")

def _pct(val):
    return f"{val:.1f}%" if val is not None else "N/A"

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════

def _sheet_summary(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = p.name
    c.font  = _font(size=16, bold=True, color=NAVY_HEX)
    c.fill  = _fill(LIGHT_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    from openpyxl.styles import numbers as num_fmt
    ORG_TYPE_LABEL = {
        "charter_school": "Charter School",
        "health_center":  "Federally Qualified Health Center",
        "early_care":     "Early Care & Education Program",
        "nonprofit":      "Nonprofit Organization",
    }
    meta = [
        ("Organization Type", ORG_TYPE_LABEL.get(p.org_type, p.org_type)),
        ("Location",          p.location),
        ("EIN",               p.ein or "—"),
        ("Year Founded",      p.year_founded or "—"),
        ("Authorizer",        p.authorizer or "N/A"),
        ("Grade Span",        p.grade_span or "N/A"),
        ("Website",           p.website or "—"),
        ("Report Date",       p.report_date or date.today().strftime("%B %d, %Y")),
        ("Prepared By",       p.prepared_by or "—"),
    ]

    row = 2
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=NAVY_HEX)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Loan block
    if p.loan:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="PROPOSED FINANCING")
        c.font  = _font(bold=True, color=WHITE_HEX, size=10)
        c.fill  = _fill(NAVY_HEX)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        loan_meta = [
            ("Loan Amount",   f"${p.loan.amount:,.0f}"),
            ("Purpose",       p.loan.purpose),
            ("Structure",     p.loan.financing_type),
            ("Term",          f"{p.loan.proposed_term_years} years" if p.loan.proposed_term_years else "—"),
            ("Collateral",    p.loan.collateral or "—"),
        ]
        for label, val in loan_meta:
            ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=NAVY_HEX)
            ws.cell(row=row, column=2, value=val)
            row += 1

    # Analyst notes
    if p.analyst_notes:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="ANALYST NOTES")
        c.font  = _font(bold=True, color=WHITE_HEX)
        c.fill  = _fill(TEAL_HEX)
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=p.analyst_notes)
        c.font = _font(italic=True)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50

    _set_col_widths(ws, [22, 40, 15, 15, 15, 15])
    return ws


def _sheet_enrollment(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Enrollment")
    ws.sheet_view.showGridLines = False

    _header_row(ws, 1, ["Year", "Enrollment", "Capacity", "Utilization %"])
    for i, e in enumerate(p.enrollment_history, start=2):
        util = (e.total / e.capacity * 100) if e.capacity else None
        _data_row(ws, i, [
            e.year, e.total, e.capacity or "—",
            f"{util:.1f}%" if util else "—"
        ], alt=(i % 2 == 0))

    _set_col_widths(ws, [12, 14, 14, 16])

    # Embedded chart image from matplotlib
    if p.enrollment_history:
        buf = charts.enrollment_chart(p.enrollment_history)
        img = XLImage(buf)
        img.width, img.height = 400, 240
        ws.add_image(img, "F2")

    return ws


def _sheet_demographics(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Demographics")
    ws.sheet_view.showGridLines = False

    if not p.demographics:
        ws["A1"] = "No demographics data provided."
        return ws

    d = p.demographics
    _header_row(ws, 1, ["Category", "Percentage"])

    race = [
        ("Black / African American", d.black),
        ("Hispanic / Latino",        d.hispanic),
        ("White",                    d.white),
        ("Asian",                    d.asian),
        ("Other / Multiracial",      d.other),
    ]
    svc = [
        ("Free / Reduced Lunch (FRL)", d.free_reduced_lunch),
        ("English Language Learners (ELL)", d.english_learners),
        ("Special Education (IEP/504)",     d.special_education),
    ]

    row = 2
    ws.cell(row=row, column=1, value="── Race / Ethnicity ──").font = _font(bold=True, color=TEAL_HEX, italic=True)
    row += 1
    for label, val in race:
        _data_row(ws, row, [label, _pct(val)], alt=(row % 2 == 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="── Service Indicators ──").font = _font(bold=True, color=TEAL_HEX, italic=True)
    row += 1
    for label, val in svc:
        _data_row(ws, row, [label, _pct(val)], alt=(row % 2 == 0))
        row += 1

    _set_col_widths(ws, [36, 18])

    # Embedded chart image
    buf = charts.demographics_chart(d)
    img = XLImage(buf)
    img.width, img.height = 420, 260
    ws.add_image(img, "D2")

    return ws


def _sheet_academics(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Academic Performance")
    ws.sheet_view.showGridLines = False

    headers = ["Year", "ELA Proficiency %", "Math Proficiency %",
               "ELA Growth (SGP)", "Math Growth (SGP)",
               "Graduation Rate %", "Attendance Rate %"]
    _header_row(ws, 1, headers)

    for i, a in enumerate(p.academic_history, start=2):
        _data_row(ws, i, [
            a.year,
            _pct(a.ela_proficiency),
            _pct(a.math_proficiency),
            _pct(a.ela_growth),
            _pct(a.math_growth),
            _pct(a.graduation_rate),
            _pct(a.attendance_rate),
        ], alt=(i % 2 == 0))

    _set_col_widths(ws, [12, 20, 20, 18, 18, 18, 18])

    if p.academic_history:
        buf = charts.academic_chart(p.academic_history)
        img = XLImage(buf)
        img.width, img.height = 460, 260
        ws.add_image(img, "I2")

    return ws


def _sheet_charter_log(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Charter Log")
    ws.sheet_view.showGridLines = False

    _header_row(ws, 1, ["Year", "Event Type", "Authorizer", "Description"])
    for i, ev in enumerate(sorted(p.charter_events, key=lambda e: e.year), start=2):
        _data_row(ws, i, [
            ev.year,
            ev.event_type.title(),
            ev.authorizer or p.authorizer or "—",
            ev.description,
        ], alt=(i % 2 == 0))
        # Colour-code event type
        color_map = {
            "original":     NAVY_HEX,
            "renewal":      GREEN_HEX,
            "modification": GOLD_HEX,
            "probation":    RED_HEX,
            "revocation":   "8B0000",
        }
        cell = ws.cell(row=i, column=2)
        cell.font = _font(bold=True, color=color_map.get(ev.event_type, GRAY_HEX))

    _set_col_widths(ws, [8, 18, 24, 60])

    if p.charter_events:
        buf = charts.charter_timeline_chart(p.charter_events)
        img = XLImage(buf)
        img.width, img.height = 520, 130
        ws.add_image(img, "A" + str(len(p.charter_events) + 4))

    return ws


def _sheet_accountability(wb, p: BorrowerProfile):
    ws = wb.create_sheet("Accountability")
    ws.sheet_view.showGridLines = False

    if not p.accountability:
        ws["A1"] = "No accountability data provided."
        return ws

    a = p.accountability
    fields = [
        ("Rating",            a.rating),
        ("Framework",         a.framework),
        ("Rating Year",       a.rating_year),
        ("Subgroup Flags",    ", ".join(a.subgroup_flags) if a.subgroup_flags else "None"),
        ("Notes",             a.notes or "—"),
    ]

    _header_row(ws, 1, ["Field", "Value"])
    for i, (label, val) in enumerate(fields, start=2):
        _data_row(ws, i, [label, val], alt=(i % 2 == 0))
        ws.cell(row=i, column=1).font = _font(bold=True, color=NAVY_HEX)

    # Colour-code rating cell
    rating_cell = ws.cell(row=2, column=2)
    rl = a.rating.lower()
    if any(w in rl for w in ["exemplary","recognized","meets","good","commended"]):
        rating_cell.font = _font(bold=True, color=GREEN_HEX, size=12)
    elif any(w in rl for w in ["improvement","probation","revoked","unsatisfactory"]):
        rating_cell.font = _font(bold=True, color=RED_HEX, size=12)
    else:
        rating_cell.font = _font(bold=True, color=GOLD_HEX, size=12)

    _set_col_widths(ws, [22, 60])
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# Main entry point
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel(profile: BorrowerProfile, output_path: str):
    wb = openpyxl.Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    _sheet_summary(wb, profile)
    if profile.enrollment_history:
        _sheet_enrollment(wb, profile)
    if profile.demographics:
        _sheet_demographics(wb, profile)
    if profile.academic_history:
        _sheet_academics(wb, profile)
    if profile.charter_events:
        _sheet_charter_log(wb, profile)
    if profile.accountability:
        _sheet_accountability(wb, profile)

    wb.save(output_path)
    return output_path
